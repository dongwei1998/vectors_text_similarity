# -*- coding:utf-8 -*-
# @Time      : 2021-04-23 16:48
# @Author    : 年少无为！
# @FileName  : file_conver.py
# @Software  : PyCharm

from config.config import *
import openpyxl
import jieba
from tqdm import tqdm
# jieba.load_userdict('D:\智慧法务项目\法律案件相似度模型\doc2vec\config\count_dice_falv.txt')

def file_conv():
    o_data = openpyxl.load_workbook(o_data_path)
    names = o_data.sheetnames
    table = o_data[names[0]]
    max_row = table.max_row
    with open('../out/fawu_text_cut.txt','w',encoding='utf-8') as w:
        with tqdm(desc="数据转换", total=max_row) as bar:
            for i in range(1,max_row):
                text = table.cell(i,12).value
                try:
                    text = ''.join(text.replace('\n','').replace(' ',''))
                except Exception as e:
                    print(e)
                    continue
                text_l = ' '.join([w for w in jieba.cut(text)])
                w.write(text_l+'\n')
                bar.update(1)

