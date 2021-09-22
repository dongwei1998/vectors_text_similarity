# -*- coding:utf-8 -*-
# @Time      : 2021-05-11 16:57
# @Author    : 年少无为！
# @FileName  : 案件输出.py
# @Software  : PyCharm
import openpyxl
o_data = openpyxl.load_workbook('../datas/text_fawu.xlsx')
names = o_data.sheetnames
table = o_data[names[0]]
max_row = table.max_row
max_column = table.max_column
with open('../similarity_result/Sheet1_all_data.txt','r',encoding='utf-8') as r:
    sim_list = r.readlines()
    for info in sim_list:   # text
        a, b, s = info.split(':')
        for i in range(1,max_row):
            if b == table.cell(i,5).value:
                table.cell(i,13).value = s
o_data.save('./20210511.xlsx')
