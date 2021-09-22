# -*- coding:utf-8 -*-
# @Time      : 2021-05-10 15:57
# @Author    : 年少无为！
# @FileName  : title_sim.py
# @Software  : PyCharm
import openpyxl
import jieba
import numpy as np
from gensim.models.fasttext import FastText
import math
from tqdm import tqdm
from gensim.models.doc2vec import Doc2Vec

# 向量合并
def voc_connect_1(word_voc):
    c = np.zeros(128, dtype=np.float32)
    n = 0
    for v in word_voc:
        c += v
        n += 1
    word_voc = list(c/n)

    return np.array(word_voc)


# 加载fasttext模型
def fasttext_lod_model(model_path):
    model_dm = FastText.load(model_path)
    return model_dm


# 获取词向量
def get_word_vecter(text,model_dm):
    voc_list = []
    for word in text:
        vector_dm = model_dm.wv.__getitem__(word)
        voc_list.append(vector_dm)
    return voc_list

# 计算相似度
def cos(final_case_vector, final_case_vector_list,top):
    array1 = final_case_vector[1]
    norm1 = math.sqrt(sum(list(map(lambda x: math.pow(x, 2), array1))))
    case_sim_list = []  # [案号，相似度]
    '''
    案号：｛案号：相似度｝
    '''
    for case in final_case_vector_list:
        array2 = case[1]
        norm2 = math.sqrt(sum(list(map(lambda x: math.pow(x, 2), array2))))
        sim = sum([array1[i] * array2[i] for i in range(0, len(array1))]) / (norm1 * norm2)
        case_sim_list.append([case[0],sim])

    case_sim_list = sorted(case_sim_list,key=lambda a:a[1],reverse=True)[:top]
    return case_sim_list

# 计算相似度
def cos_one_one(array1,array2):
    array1 = array1[1]
    norm1 = math.sqrt(sum(list(map(lambda x: math.pow(x, 2), array1))))
    norm2 = math.sqrt(sum(list(map(lambda x: math.pow(x, 2), array2))))
    sim = sum([array1[i] * array2[i] for i in range(0, len(array1))]) / (norm1 * norm2)

    return sim

# 加载doc模型
def doc_lod_model(model_path,text=('案号','我是中国人')):
    model_dm = Doc2Vec.load(model_path)
    return model_dm

# 获取文档向量
def get_text_vecter(text,model_dm):
    # test_text = [i for i in jieba.cut(text)]
    inferred_vector_dm = model_dm.infer_vector(text,steps=1000)
    return inferred_vector_dm

fast_model_path = '../models/doc2vec/fawu_doc_model'

# 加载模型
fasttext_model = doc_lod_model(fast_model_path)
# 读取数据
o_data = openpyxl.load_workbook('./20210513_sim_title.xlsx')
names = o_data.sheetnames
table = o_data[names[0]]
max_row = table.max_row
max_column = table.max_column
top = 5

for i in range(2,max_row+1):
    word_list = [word for word in jieba.cut(''.join([table.cell(i,2).value,table.cell(i,3).value,table.cell(i,4).value]))]
    word_voc = voc_connect_1(get_text_vecter(word_list, fasttext_model))
    final_case_vector_list = []
    final_case_vector = [table.cell(i,1).value,word_voc]
    with tqdm(total=(table.max_row - 2)) as pbar:
        for i1 in range(2,max_row+1):
            if i == i1:
                continue
            pbar.set_description(f'开始计算第{i}与所有案件相似度')
            word_list1 = [word for word in jieba.cut(''.join([table.cell(i1, 2).value, table.cell(i1, 3).value,table.cell(i1, 4).value]))]
            word_voc1 = voc_connect_1(get_text_vecter(word_list1, fasttext_model))
            final_case_vector_list.append([table.cell(i1, 1).value,word_voc1])
            pbar.update(1),
    final_case_top = cos(final_case_vector, final_case_vector_list,top) # [案号，相似度]
    infos = [':'.join([i[0],str(i[1])]) for i in final_case_top]
    table.cell(i, 7).value = '\n'.join(infos)
    # if i % 101 == 0:
    #     print(f'保存数据{i}条，剩余处理数据{max_row-i}条！')
    #     o_data.save('./20210512_sim_title_1.xlsx')
o_data.save('./20210513_sim_title_.xlsx')
